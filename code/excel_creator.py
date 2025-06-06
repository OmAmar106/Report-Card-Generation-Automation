import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment,Font,Border,Side
from openpyxl.drawing.image import Image
from pdfcreator import createpdf
from datetime import datetime
import os

def func1(num):
    if num==1:
        return 'I'
    elif num==2:
        return 'II'
    elif num==3:
        return 'III'
    elif num==4:
        return 'IV'
    elif num==5:
        return 'V'
    elif num==6:
        return 'VI'
    elif num==7:
        return 'VII'
    elif num==8:
        return 'VIII'
    else:
        if num in [1.5,2.5,3.5,4.5,5.5,6.5,7.5,8.5]:
            return 1
        if num in [1.25,2.25,3.25,4.25,5.25,6.25,7.25,8.25]:
            return 2
        return None

def func7(st):
    d = {"AA":10,"AB":9,"BB":8,"BC":7,"CC":6,"CD":5,"DD":4,"FF":0}
    if st in d:
        return d[st]
    elif st[-1]=='*' and st[:-1] in d:
        return d[st[:-1]]
    return 0

border_style = Border(left=Side(style="medium", color="000000"),
                      right=Side(style="medium", color="000000"),
                      top=Side(style="medium", color="000000"),
                      bottom=Side(style="medium", color="000000"))

def create(name,branch,name1,flag=True):
    df = pd.read_excel(name)
    df.columns = df.columns.str.strip()
    students = df["BT ID"].unique()

    roll_no = ''

    Failed = []
    for student in students:
        cset = set()
        if roll_no not in student:
            continue
        # y = 102
        # if count<y:
        #     count += 1
        #     continue
        # elif count==y+1:
        #     break
        try:
            student_df = df[df["BT ID"] == student].sort_values(by="Semester")

            output_filename = os.getcwd()+f"/{branch}/{'bt'+student[2:4]}"

            if not os.path.exists(output_filename):
                os.makedirs(output_filename)

            output_filename = os.getcwd()+f"/{branch}/{'bt'+student[2:4]}/{student}_Grade_Card.xlsx"
            left_align = []

            headers = []
            headers.append((2,2,2))
            headers.append((3,2,2))
            headers.append((2,5,5))
            headers.append((3,5,5))
            headers.append((2,12,2))
            headers.append((3,12,2))
            headers.append((2,15,5))
            headers.append((3,15,5))
            with pd.ExcelWriter(output_filename, engine="openpyxl") as writer:
                # header_df.to_excel(writer,sheet_name="Sheet1",startrow=0,startcol=1,index=False)
                headers.append((1,2,18))  
                left_start_row = 4
                right_start_row = 4
                
                left_start_col = 1
                right_start_col = 11
                
                tc = 0
                tacq = 0

                merge = []
                L7 = []
                table_cols = ["Course Code","Course","Credit","Grade"]
                
                unique_semesters = sorted(student_df["Semester"].unique())
                
                for semester in unique_semesters:

                    if not func1(semester):
                        continue
                    credi = 0
                    acq = 0

                    sem_df = student_df[student_df["Semester"] == semester]

                    def func(sem_df,left_start_row,left_start_col):
                        sem_df["Course Code"].to_excel(writer, sheet_name="Sheet1", startrow=left_start_row, startcol=left_start_col, index=False)
                        sem_df["Course"].to_excel(writer, sheet_name="Sheet1", startrow=left_start_row, startcol=left_start_col+2, index=False)
                        sem_df["Credit"].to_excel(writer, sheet_name="Sheet1", startrow=left_start_row, startcol=left_start_col+6, index=False)
                        sem_df["Grade"].to_excel(writer, sheet_name="Sheet1", startrow=left_start_row, startcol=left_start_col+8, index=False)

                    for index,row in sem_df.iterrows():
                        if row["Grade"]=="W":
                            continue
                        coursecode = row["Course Code"].strip()
                        credi += row["Credit"]
                        acq += row["Credit"]*func7(row["Grade"])

                        if coursecode not in cset:
                            tacq += row["Credit"]*func7(row["Grade"])
                            if "FF" not in row["Grade"]:
                                tc += row["Credit"]
                                cset.add(coursecode)
                        else:
                            tacq += row["Credit"]*func7(row["Grade"])
                        


                        # if "FF" not in row["Grade"]:
                        #     credi += row["Credit"]
                        #     tc += row["Credit"]
                        #     acq += row["Credit"]*func7(row["Grade"])
                        # else:
                        #     if  not in cset:
                        #         tc += row["Credit"]
                        #         cset.add(row["Course Code"].strip())
                        #     else:
                        #         credi += row["Credit"]

                    # tacq += acq
                    # tc += credi

                    if int(semester) % 2 == 1:
                        headers.append((left_start_row + 1, left_start_col + 1,8))
                        if semester==int(semester):
                            header_text = f"SEM. {func1(semester)} (July-Nov {int('20'+student[2:4])+int(semester)//2})"
                        else:
                            if (semester-0.25)==int(semester-0.25):
                                credi = L7[-1][-4]
                                acq += L7[-1][-2]
                                header_text = f"RE EXAM SEM. {func1(int(semester))} (July-Nov {int('20'+student[2:4])+int(semester)//2})"
                            else:
                                header_text = f"SUMMER TERM (Dec-Jan {int('20'+student[2:4])+int(semester)//2})"
                        header_df = pd.DataFrame({" ": [header_text]})
                        header_df.to_excel(writer, sheet_name="Sheet1", startrow=left_start_row, startcol=left_start_col, index=False, header=False)
                        
                        left_start_row += 1
                        func(sem_df,left_start_row,left_start_col)

                        for j in range(len(sem_df)):
                            left_align.append((left_start_row+j+2,left_start_col+3))

                        for j in range(len(sem_df)+1):
                            headers.append((left_start_row+j+1, left_start_col + 1,1))
                            headers.append((left_start_row+j+1, left_start_col + 3,3))
                            headers.append((left_start_row+j+1, left_start_col + 7,1))

                        left_start_row += len(sem_df) + 2
                        left_start_row += 1
                        merge.append((left_start_row,left_start_row+1,left_start_col+1,left_start_col+1))
                        merge.append((left_start_row,left_start_row+1,left_start_col+5,left_start_col+6))
                        L7.append((left_start_row,left_start_col,credi,tc,acq,tacq))
                        left_start_row += 2
                    
                    else:
                        headers.append((right_start_row + 1, right_start_col + 1,8))
                        if semester==int(semester):
                            header_text = f"SEM. {func1(semester)} (Jan-May {int('20'+student[2:4])+int(semester)//2})"
                        else:
                            if (semester-0.25)==int(semester-0.25):
                                credi = L7[-1][-4]
                                acq += L7[-1][-2]
                                header_text = f"RE EXAM SEM. {func1(int(semester))} (Jan-May {int('20'+student[2:4])+int(semester)//2})"
                            else:
                                header_text = f"SUMMER TERM (June-July {int('20'+student[2:4])+int(semester)//2})"
                        header_df = pd.DataFrame({" ": [header_text]})
                        header_df.to_excel(writer, sheet_name="Sheet1", startrow=right_start_row, startcol=right_start_col, index=False, header=False)
                        right_start_row += 1                    
                        func(sem_df,right_start_row,right_start_col)

                        for j in range(len(sem_df)):
                            left_align.append((right_start_row+j+2,right_start_col+3))

                        for j in range(len(sem_df)+1):
                            headers.append((right_start_row+j+1, right_start_col + 1,1))
                            headers.append((right_start_row+j+1, right_start_col + 3,3))
                            headers.append((right_start_row+j+1, right_start_col + 7,1))
                        right_start_row += len(sem_df) + 2
                        right_start_row += 1
                        merge.append((right_start_row,right_start_row+1,right_start_col+1,right_start_col+1))
                        merge.append((right_start_row,right_start_row+1,right_start_col+5,right_start_col+6))
                        L7.append((right_start_row,right_start_col,credi,tc,acq,tacq))
                        right_start_row += 2

            wb = load_workbook(output_filename)
            ws = wb.active

            ws.column_dimensions['A'].width = 1
            ws.column_dimensions['K'].width = 3
            # column_widths = {"A": 20, "B": 30, "C": 15}

            # for col, width in column_widths.items():
                # ws.column_dimensions[col].width = width
            # column_widths = {"C": 30, "H": 30, "B": 15,"G":15}

            bold_font = Font(bold=True)
            for row, col,k in headers:
                ws.cell(row=row, column=col).font = bold_font
                ws.merge_cells(start_row=row,end_row=row,start_column=col,end_column=col+k)
            
            for u,v,w,x in merge:
                ws.merge_cells(start_row=u,end_row=v,start_column=w,end_column=x)

            ws.row_dimensions[1].height = 169
            if flag:
                img = Image(os.path.dirname(os.path.abspath(__file__))+'\\data\\IIITN-Logo.jpg')
                img.height = 240
                img.width = 1221 # update this accorindg to the height
                ws.add_image(img,'B1')
                
            # for col, width in column_widths.items():
            #     ws.column_dimensions[col].width = width

            k = ws["B2"]
            k.value = "Name : " 
            k.font = Font(bold=True,size=13)
            k = ws["E2"]
            k.value = name1[student] 
            k.font = Font(bold=True,size=13)
            k = ws["B3"]
            k.value = "Branch : "
            k.font = Font(bold=True,size=13)
            k = ws["E3"]
            k.value = branch
            k.font = Font(bold=True,size=13)

            k = ws["L2"]
            k.value = "Enrollment No. : " 
            k.font = Font(bold=True,size=13)
            k = ws["O2"]
            k.value = student.upper() 
            k.font = Font(bold=True,size=13)
            k = ws["L3"]
            k.value = "Degree : "
            k.font = Font(bold=True,size=13)
            k = ws["O3"]
            k.value = "Bachelor of Technology"
            k.font = Font(bold=True,size=13)
            
            for startrow,startcol,credits,totalcredits,egp,totalegp in L7:
                k = ws.cell(startrow,startcol+1)
                k.value = "SGPA"

                k = ws.cell(startrow,startcol+2)
                k.value = "Credit"
                k = ws.cell(startrow+1,startcol+2)
                k.value = credits

                k = ws.cell(startrow,startcol+3)
                k.value = "EGP"
                k = ws.cell(startrow+1,startcol+3)
                k.value = egp

                k = ws.cell(startrow,startcol+4)
                k.value = "SGPA"
                k = ws.cell(startrow+1,startcol+4)
                if credits!=0:
                    k.value = "%.2f"%(egp/credits)
                else:
                    k.value = 0

                k = ws.cell(startrow,startcol+5)
                k.value = "CGPA"

                k = ws.cell(startrow,startcol+7)
                k.value = "Credit"
                k = ws.cell(startrow+1,startcol+7)
                k.value = totalcredits

                k = ws.cell(startrow,startcol+8)
                k.value = "EGP"
                k = ws.cell(startrow+1,startcol+8)
                k.value = totalegp

                k = ws.cell(startrow,startcol+9)
                k.value = "CGPA"
                k = ws.cell(startrow+1,startcol+9)
                k.value = "%.2f"%(totalegp/totalcredits)

            for row in ws.iter_rows():
                max_height=20
                for cell in row:
                    if isinstance(cell.value, str) or isinstance(cell.value, int):
                        # Exclude B1 from having a border
                        if cell.coordinate != "B1":
                            cell.border = border_style
                    if cell.coordinate in ['L3', 'O2', 'L2', 'O3', 'B2', 'E2', 'B3', 'E3']:
                        continue
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.font = Font(bold=True)
                    # if cell.value:
                    #     txt_len=len(str(cell.value))
                    #     if(txt_len>35):
                    #         row_max=15+(txt_len//35)*10
                    #         try:
                    #             ws.row_dimensions[cell.row].height=max(ws.row_dimensions[cell.row].height,row_max)
                    #         except:
                    #             ws.row_dimensions[cell.row].height= row_max
            
            for u,v in left_align:
                cell = ws.cell(row=u, column=v)
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                cell.font = Font(bold=True)
                if cell.value:
                    txt_len=len(str(cell.value))
                    if(txt_len>40):
                        row_max=20+(txt_len//35)*10
                        try:
                            ws.row_dimensions[cell.row].height=max(ws.row_dimensions[cell.row].height,row_max)
                        except:
                            ws.row_dimensions[cell.row].height= row_max
                
            # cell = ws["B1"]
            # cell.value = "GRADE CARD"
            # cell.font = Font(bold=True,size=20)
            ws.merged_cells.ranges = [range for range in ws.merged_cells.ranges if not (range.min_row == 1 and range.min_col == 2)]
            for merged_range in ws.merged_cells.ranges:
                    min_row, min_col, max_row, max_col = merged_range.min_row, merged_range.min_col, merged_range.max_row, merged_range.max_col
                    for row in range(min_row, max_row + 1):
                        for col in range(min_col, max_col + 1):
                            ws.cell(row=row, column=col).border = border_style

            # ws.cell(row=6, column=4).alignment = Alignment(horizontal='left',vertical='top')
            if flag:
                t = max(right_start_row,left_start_row)+1
                t += 1
                ws.cell(row=t, column=2).value = "Note: This grade card is exclusively for internal use"
                ws.cell(row=t, column=2).font = Font(bold=True,size=13)
                
                t += 1
                ws.cell(row=t, column=2).value = "Medium of Instruction :English"
                ws.cell(row=t, column=2).font = Font(bold=True,size=13)

                t += 1
                ws.cell(row=t, column=2).value = "Abreviations: SGPA-Semester Grade Point Average, CGPA:-Cumulative Grade Point Average, EGP-Earned Grade Points"
                ws.cell(row=t, column=2).font = Font(bold=True,size=13)

                t += 1
                ws.cell(row=t, column=2).value = "(The statement is subject to correction, if any)"
                ws.cell(row=t, column=2).font = Font(bold=True,size=13)

                t += 1
                ws.cell(row=t, column=2).value = "Date: "+datetime.today().strftime("%d.%m.%Y")
                ws.cell(row=t, column=2).font = Font(bold=True,size=13)

                t += 1
                ws.cell(row=t, column=2).value = "THIS IS ELECTRONICALLY GENERATED DOCUMENT AND DOES NOT REQUIRE SIGNATURE"
                ws.cell(row=t, column=2).font = Font(bold=True,size=13)

            wb.save(output_filename)
            wb.close()

            # print(f"Created file for {student}: {output_filename}")
            createpdf(output_filename)
            os.remove(output_filename)
            # count += 1
            print("Processed Student : "+student+" ,Total Failed : "+str(len(Failed)))
        except:
            print("Failed Student : "+student+" ,Total Failed : "+str(len(Failed)))
            Failed.append(student)
        # break
            
    print("Processing complete.")
    print("Failed for : ",Failed)
